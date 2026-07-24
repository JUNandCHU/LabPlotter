from __future__ import annotations

import csv
import re
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path
from typing import Any, Callable, Iterable

import numpy as np
from openpyxl import load_workbook

from .i18n import tr
from .models import Spectrum, ZetaMeasurement


SS_NS = "urn:schemas-microsoft-com:office:spreadsheet"
NS = {"s": SS_NS}


def _coerce_pair(a: Any, b: Any) -> tuple[float, float] | None:
    try:
        x, y = float(a), float(b)
    except (TypeError, ValueError):
        return None
    if not (np.isfinite(x) and np.isfinite(y)):
        return None
    return x, y


def _delimited_rows(path: str | Path) -> list[list[str]]:
    text = Path(path).read_text(encoding="utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        return list(csv.reader(text.splitlines(), dialect))
    except csv.Error:
        return [re.split(r"[,;\t\s]+", line.strip()) for line in text.splitlines()]


def parse_ftir_file(path: str | Path) -> Spectrum:
    path = Path(path)
    suffix = path.suffix.lower()
    pairs: list[tuple[float, float]] = []
    if suffix in {".csv", ".txt", ".tsv"}:
        for row in _delimited_rows(path):
            if len(row) >= 2 and (pair := _coerce_pair(row[0], row[1])):
                pairs.append(pair)
    elif suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(values_only=True):
            if len(row) >= 2 and (pair := _coerce_pair(row[0], row[1])):
                pairs.append(pair)
    else:
        raise ValueError(tr("Unsupported FTIR file: {suffix}", suffix=path.suffix))
    if len(pairs) < 3:
        raise ValueError(tr("Could not find at least three numeric X/Y rows."))
    arr = np.asarray(pairs, dtype=float)
    return Spectrum(path.stem, arr[:, 0], arr[:, 1], str(path), metadata={"kind": "ftir"}).clean()


def _xml_row_values(row: ET.Element) -> list[Any]:
    values: list[Any] = []
    col = 1
    for cell in row.findall("s:Cell", NS):
        index = cell.attrib.get(f"{{{SS_NS}}}Index")
        if index:
            col = int(index)
        while len(values) < col - 1:
            values.append(None)
        data = cell.find("s:Data", NS)
        value: Any = data.text if data is not None else None
        if data is not None and data.attrib.get(f"{{{SS_NS}}}Type") == "Number":
            try:
                value = float(value)
            except (TypeError, ValueError):
                pass
        values.append(value)
        col += 1
    return values


def parse_nanodrop_xml(path: str | Path) -> list[Spectrum]:
    path = Path(path)
    root = ET.parse(path).getroot()
    spectra: list[Spectrum] = []
    for ws in root.findall("s:Worksheet", NS):
        sheet_name = ws.attrib.get(f"{{{SS_NS}}}Name", "Sheet")
        rows = [_xml_row_values(row) for row in ws.findall("s:Table/s:Row", NS)]
        if len(rows) < 3:
            continue
        title = str(rows[1][2]).strip() if len(rows[1]) > 2 and rows[1][2] else sheet_name
        measured_at = rows[1][3] if len(rows[1]) > 3 else ""
        pairs = [p for row in rows[1:] if len(row) >= 2 and (p := _coerce_pair(row[0], row[1]))]
        if len(pairs) < 3:
            continue
        arr = np.asarray(pairs)
        spectra.append(
            Spectrum(
                title,
                arr[:, 0],
                arr[:, 1],
                str(path),
                visible="blank" not in title.lower(),
                metadata={"kind": "nanodrop", "sheet": sheet_name, "blank": "blank" in title.lower(), "time": measured_at},
            ).clean()
        )
    return spectra


def parse_nanodrop_xlsx(path: str | Path) -> list[Spectrum]:
    path = Path(path)
    wb = load_workbook(path, data_only=True, read_only=True)
    spectra: list[Spectrum] = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            continue
        title = str(rows[1][2]).strip() if len(rows[1]) > 2 and rows[1][2] else ws.title
        measured_at = rows[1][3] if len(rows[1]) > 3 else ""
        pairs = [p for row in rows[1:] if len(row) >= 2 and (p := _coerce_pair(row[0], row[1]))]
        if len(pairs) < 3:
            continue
        arr = np.asarray(pairs)
        spectra.append(Spectrum(title, arr[:, 0], arr[:, 1], str(path), visible="blank" not in title.lower(), metadata={"kind": "nanodrop", "sheet": ws.title, "blank": "blank" in title.lower(), "time": str(measured_at)}).clean())
    return spectra


def parse_nanodrop_file(path: str | Path) -> list[Spectrum]:
    suffix = Path(path).suffix.lower()
    if suffix == ".xml":
        return parse_nanodrop_xml(path)
    if suffix in {".xlsx", ".xlsm"}:
        return parse_nanodrop_xlsx(path)
    raise ValueError(tr("Unsupported NanoDrop file: {suffix}", suffix=suffix))


def _sample_from_header(header: str, sheet_name: str) -> str:
    match = re.search(r"\s-\s(.+?)(?:\s*\[|$)", header)
    if match:
        return match.group(1).strip()
    cleaned = re.sub(r"(?i)_?(size|zeta)(?:_cell_?\d+)?$", "", sheet_name).strip(" _-")
    return cleaned or sheet_name


def parse_zetasizer_workbook(
    path: str | Path,
    progress_callback: Callable[[int, int, str], None] | None = None,
) -> list[ZetaMeasurement]:
    path = Path(path)
    wb = load_workbook(path, data_only=True, read_only=False)
    output: list[ZetaMeasurement] = []
    total_sheets = max(1, len(wb.worksheets))
    for sheet_index, ws in enumerate(wb.worksheets, start=1):
        if progress_callback is not None:
            progress_callback(sheet_index, total_sheets, ws.title)
        header = str(ws.cell(1, 1).value or "")
        lower = header.lower()
        if "size (d.nm)" in lower:
            kind = "DLS"
        elif "zeta potential" in lower:
            kind = "Zeta"
        else:
            continue
        particle = _sample_from_header(header, ws.title)
        cell_match = re.search(r"(?i)cell[_\s-]*(\d+)", ws.title)
        cell_number = cell_match.group(1) if cell_match else ""
        images = sorted(getattr(ws, "_images", []), key=lambda image: image.anchor._from.row)
        for replicate, x_col in enumerate((1, 3, 5), start=1):
            pairs = []
            for row in range(2, ws.max_row + 1):
                pair = _coerce_pair(ws.cell(row, x_col).value, ws.cell(row, x_col + 1).value)
                if pair:
                    pairs.append(pair)
            if len(pairs) < 3:
                continue
            arr = np.asarray(pairs)
            result_png = images[replicate - 1]._data() if replicate <= len(images) else None
            output.append(
                ZetaMeasurement(
                    particle,
                    kind,
                    replicate,
                    arr[:, 0],
                    arr[:, 1],
                    str(path),
                    ws.title,
                    cell_number,
                    result_png,
                    {"header_x": header, "header_y": str(ws.cell(1, x_col + 1).value or "")},
                )
            )
    if not output:
        raise ValueError(tr("No DLS or zeta-potential raw data sheets were detected."))
    return output


def workbook_preview(path: str | Path, limit_rows: int = 30, limit_cols: int = 12) -> dict[str, list[list[Any]]]:
    """Return small sheet previews for the custom format mapper."""
    path = Path(path)
    if path.suffix.lower() in {".csv", ".tsv", ".txt"}:
        return {path.stem: [row[:limit_cols] for row in _delimited_rows(path)[:limit_rows]]}
    if path.suffix.lower() == ".xml":
        root = ET.parse(path).getroot()
        return {
            ws.attrib.get(f"{{{SS_NS}}}Name", "Sheet"): [_xml_row_values(row)[:limit_cols] for row in ws.findall("s:Table/s:Row", NS)[:limit_rows]]
            for ws in root.findall("s:Worksheet", NS)
        }
    wb = load_workbook(path, data_only=True, read_only=True)
    previews: dict[str, list[list[Any]]] = {}
    for ws in wb.worksheets:
        max_row = min(limit_rows, ws.max_row or limit_rows)
        max_col = min(limit_cols, ws.max_column or limit_cols)
        previews[ws.title] = [list(row) for row in ws.iter_rows(values_only=True, max_row=max_row, max_col=max_col)]
    return previews


def workbook_signature(path: str | Path) -> dict[str, Any]:
    """A stable, non-data-value fingerprint for matching saved format profiles."""
    path = Path(path)
    previews = workbook_preview(path, limit_rows=8, limit_cols=24)
    first_rows = []
    for rows in previews.values():
        row = next((r for r in rows if any(v not in (None, "") for v in r)), [])
        tokens = []
        for value in row:
            if value in (None, ""):
                tokens.append("")
            elif isinstance(value, (int, float)):
                tokens.append("<number>")
            else:
                text = str(value).strip().lower()
                try:
                    float(text)
                    tokens.append("<number>")
                except ValueError:
                    tokens.append(re.sub(r"\s+", " ", text)[:80])
        first_rows.append(tokens)
    return {"extension": path.suffix.lower(), "sheet_count": len(previews), "first_rows": first_rows[:3]}


def detect_builtin_kind(path: str | Path) -> str | None:
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".zip":
        try:
            with zipfile.ZipFile(path) as archive:
                names = archive.namelist()
            if any(name.endswith("/acqus") for name in names) and any(name.endswith(("/fid", "/ser")) for name in names):
                return "ssNMR"
        except (OSError, zipfile.BadZipFile):
            return None
    if suffix in {".csv", ".tsv", ".txt"}:
        return "FTIR"
    if suffix == ".xml":
        preview = workbook_preview(path, limit_rows=2, limit_cols=4)
        row = next(iter(preview.values()), [[]])[0]
        text = " ".join(str(v or "") for v in row).lower()
        return "NanoDrop" if "wavelength" in text and "absorbance" in text else None
    if suffix in {".xlsx", ".xlsm"}:
        preview = workbook_preview(path, limit_rows=2, limit_cols=6)
        headers = [" ".join(str(v or "") for v in rows[0]).lower() for rows in preview.values() if rows]
        if any("size (d.nm)" in h or "zeta potential" in h for h in headers):
            return "ZetaSizer"
        if any("wavelength" in h and "absorbance" in h for h in headers):
            return "NanoDrop"
    return None


def column_index(value: str | int) -> int:
    if isinstance(value, int):
        return value
    text = value.strip().upper()
    if text.isdigit():
        return int(text) - 1
    result = 0
    for char in text:
        if not "A" <= char <= "Z":
            raise ValueError(tr("Invalid column: {column}", column=value))
        result = result * 26 + ord(char) - ord("A") + 1
    return result - 1


def parse_generic_with_profile(path: str | Path, profile: dict[str, Any]) -> list[Spectrum]:
    previews = workbook_preview(path, limit_rows=1_000_000, limit_cols=256)
    sheet = profile.get("sheet") or next(iter(previews))
    if sheet not in previews:
        names = list(previews)
        index = int(profile.get("sheet_index", 0))
        if not 0 <= index < len(names):
            raise ValueError(tr("Sheet '{sheet}' is not present.", sheet=sheet))
        sheet = names[index]
    rows = previews[sheet]
    start = max(0, int(profile.get("data_start_row", 2)) - 1)
    header_row = max(0, int(profile.get("header_row", 1)) - 1)
    x_col = column_index(profile.get("x_column", "A"))
    y_cols = [column_index(v) for v in profile.get("y_columns", ["B"])]
    result = []
    for y_col in y_cols:
        pairs = []
        for row in rows[start:]:
            if len(row) > max(x_col, y_col) and (pair := _coerce_pair(row[x_col], row[y_col])):
                pairs.append(pair)
        if len(pairs) < 3:
            continue
        header = rows[header_row][y_col] if header_row < len(rows) and y_col < len(rows[header_row]) else None
        name = str(header or f"{Path(path).stem} Col {y_col + 1}")
        arr = np.asarray(pairs)
        result.append(Spectrum(name, arr[:, 0], arr[:, 1], str(path), metadata={"kind": "generic", "profile": profile.get("name", "")}).clean())
    if not result:
        raise ValueError(tr("The selected mapping did not produce numeric X/Y data."))
    return result
