from __future__ import annotations

from io import BytesIO
import unittest

import numpy as np
from openpyxl import Workbook
from PIL import Image

from labplotter.plotting import PlotOptions, figure_png_bytes
from labplotter.tem import TEMAnalysisParameters
from labplotter.web import (
    parse_uploaded_payload,
    processed_ftir_spectra,
    spectra_figure,
    tem_distribution_figure,
    zetasizer_curve_figure,
    zetasizer_peak_summary,
)


class WebCoreTests(unittest.TestCase):
    def test_browser_payload_uses_shared_ftir_parser_and_plotter(self):
        payload = b"Wavenumber,Transmittance\n4000,100\n3000,90\n2000,80\n1000,95\n"
        result = parse_uploaded_payload("../example.csv", payload, "FTIR")
        self.assertEqual(result.spectra[0].source, "example.csv")
        processed = processed_ftir_spectra(result.spectra, normalization_enabled=True)
        figure = spectra_figure(
            processed,
            PlotOptions("Wavenumber", "cm^-1", "Transmittance", "a.u.", font_family="DejaVu Sans", reverse_x=True),
        )
        self.assertTrue(figure_png_bytes(figure).startswith(b"\x89PNG"))
        self.assertGreater(figure.axes[0].get_xlim()[0], figure.axes[0].get_xlim()[1])

    def test_browser_payload_uses_shared_zetasizer_parser_and_figures(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "JM10A_Size"
        for x_col in (1, 3, 5):
            sheet.cell(1, x_col, "Size (d.nm) - JM10A")
            sheet.cell(1, x_col + 1, "Intensity (%)")
            for row, (x, y) in enumerate(((10, 1), (100, 5), (1000, 2)), start=2):
                sheet.cell(row, x_col, x)
                sheet.cell(row, x_col + 1, y)
        stream = BytesIO()
        workbook.save(stream)
        result = parse_uploaded_payload("zeta.xlsx", stream.getvalue(), "ZetaSizer")
        self.assertEqual(len(result.measurements), 3)
        rows = zetasizer_peak_summary(result.measurements, "DLS")
        self.assertEqual(rows[0]["n"], 3)
        self.assertEqual(rows[0]["Mean peak"], 100)
        figure = zetasizer_curve_figure(
            result.measurements,
            "DLS",
            ["JM10A"],
            PlotOptions("Particle diameter", "nm", "Intensity", "%"),
        )
        self.assertEqual(figure.axes[0].get_xscale(), "log")

    def test_browser_tem_analysis_does_not_require_persistent_library(self):
        image = np.full((512, 512), 230, dtype=np.uint8)
        yy, xx = np.ogrid[:512, :512]
        image[(xx - 180) ** 2 + (yy - 180) ** 2 <= 45 ** 2] = 60
        image[475:479, 30:130] = 254
        stream = BytesIO()
        Image.fromarray(image, mode="L").save(stream, format="TIFF")
        result = parse_uploaded_payload(
            "Batch_A_50000X_0001.tif",
            stream.getvalue(),
            "TEM",
            tem_parameters=TEMAnalysisParameters(maximum_diameter_nm=500),
        )
        self.assertEqual(result.tem.batch_name, "Batch_A")
        self.assertEqual(result.tem.status, "analyzed")
        self.assertGreater(result.tem.particle_count, 0)
        figure = tem_distribution_figure([result.tem])
        self.assertTrue(figure_png_bytes(figure).startswith(b"\x89PNG"))


if __name__ == "__main__":
    unittest.main()
